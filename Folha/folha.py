import openpyxl
import calendar
from datetime import datetime
import requests
import json
import win32com.client as win32

wb = openpyxl.load_workbook(r'C:\Users\loliveira\PycharmProjects\Automacao-de-Tarefas\Automacao-de-Tarefas\Folha\Folha modelo.xlsx')  # Abrir arquivo com a base
sheets = wb.sheetnames
ws = wb['Plan1']

mes = f'{str(datetime.today().month - 1).zfill(2)}/{str(datetime.today().year)}'

ws['A2'] = f'MÊS e ANO: {mes}'

# Formatar data a partir do mês desejado e criar range do primeiro ao último dia do mês
data = datetime.strptime(mes, '%m/%Y')
final_mes = calendar.monthrange(data.year, data.month)[1]
dias = [datetime(data.year, data.month, dia) for dia in range(1, final_mes+1)]

# Buscando feriados do mês pela API pela brasilapi
url = f'https://brasilapi.com.br/api/feriados/v1/{data.year}'
r = requests.get(url, timeout=None)
dados_feriados = json.loads(r.content)

# adicionar feriados a uma lista
feriados_mes = [datetime.strptime(i['date'], '%Y-%m-%d').day for i in dados_feriados if
                datetime.strptime(i['date'], '%Y-%m-%d').month == data.month]

# Adicionar dias que sofrem alterações na planilha e que podem não estar no mês anterior
ws['I18'] = 29
ws['I19'] = 30
ws['I20'] = 31

qtde_dias = len(dias)

"""
O preenchimento básico da folha deve ser feito utilizando os seguintes códigos:
P -Presença
R -Repouso Semanal ou Feriado
C -Dia Ponte Compensado

Criar um dicionário com os dias do mês, utilizando como chave os codigos descritos acima
"""

dados = {}
for ind, i in enumerate(dias):
    if i.day not in feriados_mes:
        if i.weekday() < 5:
            dados.update({i.day: 'P'})
        else:
            dados.update({i.day: 'R'})
    else:
        if i.weekday() + 1 == 4:
            dados.update({i.day: 'R'})
            dados.update({i.day+1: 'C'})
            dias.pop(ind)
        elif i.weekday() - 1 == 0:
            dados.update({i.day: 'R'})
            dados.update({i.day - 1: 'C'})
        else:
            dados.update({i.day: 'R'})


indice_linha = 1
indice_coluna = 2
for i in range(3):  # Range igual a 3 para alcançar as 3 colunas da planilha com datas
    for row in ws.iter_rows(min_row=12, min_col=indice_coluna, max_col=indice_coluna, max_row=22):
        for cell in row:  # Iterar sobre as colunas para preencher o codigo correspondente
            if cell.value == 'USO EXCLUSIVO DO RH' or cell.value == 'None':
                indice_linha += 1
                break
            else:
                cell.value = dados.get(indice_linha)
                indice_linha += 1
    indice_coluna += 4

# Formatar a planilha de acordo com a qtde de dias do mês, apagando os dias que não existem
if qtde_dias == 28:
    ws['I18'] = '-'
    ws['I19'] = '-'
    ws['I20'] = '-'
    ws['J18'] = '-'
    ws['J19'] = '-'
    ws['J20'] = '-'
elif qtde_dias == 29:
    ws['I19'] = '-'
    ws['I20'] = '-'
    ws['J19'] = '-'
    ws['J20'] = '-'
elif qtde_dias == 30:
    ws['I20'] = '-'
    ws['J20'] = '-'


wb.save(r'C:\Users\loliveira\PycharmProjects\Automacao-de-Tarefas\Automacao-de-Tarefas\Folha\Folha.xlsx')

outlook = win32.Dispatch('outlook.application')
# criar um email
email = outlook.CreateItem(0)
# configurar as informações do e-mail e selecionar o endereço pelo arquivo de texto
email.To = 'loliveira@gasbrasiliano.com.br'
email.Subject = "Folha de Presença"
email.HTMLBody = f"""
                   <p>Folha de Presença de {mes} disponível para validação.</p>

                   """
email.Send()


